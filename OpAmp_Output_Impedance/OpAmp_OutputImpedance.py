from decimal import Decimal
import chromedriver_autoinstaller
import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import zipfile
import json
import time
import os
import shutil
from datetime import datetime
import unittest
import math
import PySpice
import ltspice
import PyLTSpice
from PyLTSpice import SimCommander
from PySpice.Spice.Library import SpiceLibrary
from PySpice.Spice.Netlist import Circuit
from PySpice.Unit import *
from openpyxl import load_workbook
from openpyxl.chart import  (ScatterChart, Reference, Series)
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import range_boundaries
import sympy
import re
import glob


class OpAmp(unittest.TestCase):

    def setUp(self):
        #driver instance
        options = Options()
        options.add_argument("--headless=new")
        chromedriver_autoinstaller.install()
        self.driver = webdriver.Chrome(options=options)
        with open(r'opAmp_OutputImpedance.json')as d:
            self.testData = json.load(d)['Variables'][0]

    def test_export(self):
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        print("        opAmp_OutputImpedance script is running...       ")
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        driver = self.driver
        driver.maximize_window()
        driver.get(self.testData['URL'])

        #Accept Cookies
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#noise-spinner")))
        WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.CSS_SELECTOR, "#noise-spinner")))
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.consent-dialog:nth-child(1) div.modal.fade.in.show "
                            "div.modal-dialog div.modal-content div.modal-body div.short-description > a.btn.btn-success:nth-child(2)"))).click()

        gain = self.testData['gain']
        device = self.testData['device']
        R2 = self.testData['R2']
        C2VALUE = self.testData['C2VALUE']

        #Run the simulation in Nimble
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/div[@id='build-signal-chain-tab-content']"
            "/div[@id='adi-signal-chain-row']/div[@id='analog-signal-chain-group']/div[@id='signal-chain-drop-area']/table[1]/tr[1]/td[1]/div[1]/div[2]/div[2]/div[1]/*[1]" ))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(Keys.CONTROL + "a")
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(Keys.DELETE)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(gain)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CSS_SELECTOR, "#text6747-2-6 > tspan.schematic-edit-icon.schematic-part-edit-selection-link.schematic-edit-selection-link" ))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#filter-0'))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#filter-0'))).send_keys(device)
        #WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#partSelectModal2-title"))).click()

        #check if part is present or not disabled in Nimble list
        try:
            element = driver.find_element(By.CSS_SELECTOR, "#device-table > div.slick-pane.slick-pane-top.slick-pane-left > div.slick-viewport.slick-viewport-top.slick-viewport-left > div > div")
            class_attribute = element.get_attribute('class')
            if class_attribute and 'disabled' in class_attribute:
                raise Exception(device + " can't be selected in Nimble list")
            else:
                element.click()
        except NoSuchElementException:
            raise Exception(device + " can't be selected in Nimble list")
            
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CSS_SELECTOR, 'body.ember-application.modal-open:nth-child(2) div.adi-modal.modal-fills-window.modal-hide-scroll:nth-child(5) div.modal.fade.show.d-block:nth-child(1) '
            'div.modal-dialog div.modal-content div.modal-body div.configure-amp.configure-signal-chain-item div.adi-modal.modal-fills-window:nth-child(5) '
            'div.modal.fade.show.d-block:nth-child(1) div.modal-dialog div.modal-content div.modal-footer div.button-row > button.btn.btn-primary:nth-child(1)' ))).click()

        # Dictionary converting kilo, Mega
        d = {'k': 1000, 'M': 1000000, 'f': 1e-15, 'p': 1e-12, 'n': 1e-9, 'u': 1e-6}
        def text_to_num(text):
            if text[-1] in d:
                num, magnitude = text[:-1], text[-1]
                return float(num) * d[magnitude]
            else:
                return Decimal(text)
        new_rvalue = text_to_num(R2)
        new_c2value = text_to_num(C2VALUE)
  
        #This function gets the Slider Value to be passed by javascript command     
        def value_to_position(value, limit1, limit2):
            minpos = 1
            maxpos = 10000
            minval = math.log(limit1)
            maxval = math.log(limit2)
            scale = (maxval - minval) / (maxpos - minpos)

            if value <= 0: 
                return minpos 
            else: 
                position = minpos + (math.log(value) - minval) / scale 
                return position
                       
        rposition = value_to_position(new_rvalue, 10, 10000000)
        c2position = value_to_position(new_c2value, 1e-15, 1e-6)

        if (float(gain)!= 1):
            driver.execute_script(f"document.querySelector('#rscale-slider').value = {rposition}; document.querySelector('#rscale-slider').dispatchEvent(new Event('input'));")
            driver.execute_script(f"document.querySelector('#c2-slider').value = {c2position}; document.querySelector('#c2-slider').dispatchEvent(new Event('input'));")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("                    Slider values set!                   ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        else:
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("                     No Slider values                    ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        time.sleep(2)

        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//*[@id=\"config-signal-chain-item-modal\"]/div[1]/div/div/div[3]/div/button[1]"))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/"
                      "div[@id='next-steps-tab-content']/div[@id='next-steps-container']/div[2]/div[1]"))).click()
        time.sleep(5)

        #This script is extracting files directly to project folder
        now = datetime.now()
        day = str(now.day)
        current_date = now.strftime(f"%B {day}, %Y")
        downloads_path = self.testData['downloads_path']
        project_path = self.testData['project_location']
        file_path = downloads_path + 'Full Data Export - ' + current_date + '.zip'
        with zipfile.ZipFile(file_path) as zip_ref:
            new_path = project_path + '\\' + device + ' extracted files'
            zip_ref.extractall(new_path)
        print("Files were extracted to project folder")

        #Deletes the zip file after extracting
        if os.path.exists(file_path):
            os.remove(file_path)
        else:
            print("Zip file does not exist")  

        # Running the simulation in LTSpice          
        file_path = project_path + '\\' + device + ' extracted files' + '\\' + 'Ltspice Schematics'

        # netlists are created
        LTC = SimCommander(file_path + "\\AC_Simulation.asc")

        #changing netlist file into txt file
        old_path = (file_path + "\\AC_Simulation.net")
        new_path = (file_path + "\\AC_Simulation_Result.txt")
        
        dir_path, filename = os.path.split(old_path)
        name, ext = os.path.splitext(filename)
        new_name, new_ext = os.path.splitext(new_path)
        new_filename = new_name + new_ext
        new_path = os.path.join(dir_path, new_filename)
        shutil.copy(old_path, new_path)

        #making changes in the txt file
        with open(new_path, "r") as f:
            lines = f.readlines()
        del lines[1]
        
        if int(gain) == 1:
            lines = [line.replace("N002", "N001") for line in lines]
            second_occurrence_index = lines[1].find('N001', lines[1].find('N001') + 1)
            lines[1] = lines[1][:second_occurrence_index] + '0' + lines[1][second_occurrence_index + len('N001'):]
            lines.insert(6, "I1 0 out 0 AC 1\n")
            if device not in lines[3]:
                raise Exception('LTspice data not available for ' + device )
        elif int(gain) > 0:
            lines = [line.replace("N003", "N002") for line in lines]  
            second_occurrence_index = lines[1].find('N002', lines[1].find('N002') + 1)
            lines[1] = lines[1][:second_occurrence_index] + '0' + lines[1][second_occurrence_index + len('N002'):]
            lines.insert(8, "I1 0 out 0 AC 1\n")
            if device not in lines[3]:
                raise Exception('LTspice data not available for ' + device )
        elif int(gain) < 0:
            lines = [line.replace("N002", "N001") for line in lines] 
            lines = [line.replace("N003", "N002") for line in lines]
            second_occurrence_index = lines[1].find('N001', lines[1].find('N001') + 1)
            lines[1] = lines[1][:second_occurrence_index] + '0' + lines[1][second_occurrence_index + len('N001'):]
            lines.insert(8, "I1 0 out 0 AC 1\n")
            if device not in lines[3]:
                raise Exception('LTspice data not available for ' + device )

        #extract values of VDD-1 and VSS-1 and store them in variables
        if int(gain) == 1:
            line5 = lines[4].rstrip()  # remove trailing newline character
            num1_str = line5.split()[-1]  # extract last space-separated element of line
            sym1 = float(num1_str)  # convert string to float

            line6 = lines[5].rstrip()  # remove trailing newline character
            num2_str = line6.split()[-1]  # extract last space-separated element of line
            sym2 = float(num2_str)  # convert string to float
        else:
            line7 = lines[6].rstrip()  # remove trailing newline character
            num1_str = line7.split()[-1]  # extract last space-separated element of line
            sym1 = float(num1_str)  # convert string to float

            line8 = lines[7].rstrip()  # remove trailing newline character
            num2_str = line8.split()[-1]  # extract last space-separated element of line
            sym2 = float(num2_str)  # convert string to float



        #making numbers symmetrical
        def symmetrical(num1, num2):
            abs_num1 = abs(num1)
            abs_num2 = abs(num2)

            if abs_num1 == abs_num2:
                return num1, num2
            
            avg = (abs_num1 + abs_num2) / 2
            sym_abs_num1 = 2 * avg - abs_num1
            sym_abs_num2 = 2 * avg - abs_num2
            
            if num1 < 0:
                sym_num1 = -sym_abs_num1
            else:
                sym_num1 = sym_abs_num1
            if num2 < 0:
                sym_num2 = -sym_abs_num2
            else:
                sym_num2 = sym_abs_num2
            if sym_num1 == sym_num2:
                return sym_num1, sym_num2
            else:
                return (num1 + num2) / 2, -(num1 + num2) / 2
            
        num1 = sym1
        num2 = sym2
        sym_num1, sym_num2 = symmetrical(num1, num2)

        if self.testData['gain'] == '1':
            line5_new = line5.replace(num1_str, str(sym_num1))  # replace last number with sym_num1
            lines[4] = line5_new + "\n"  # add newline character back and update the list

            line6_new = lines[5].rsplit(' ', 1)[0] + f' {sym_num2}\n'
            lines[5] = line6_new
        else:
            line7_new = line7.replace(num1_str, str(sym_num1))  # replace last number with sym_num1
            lines[6] = line7_new + "\n"  # add newline character back and update the list

            line8_new = lines[7].rsplit(' ', 1)[0] + f' {sym_num2}\n'
            lines[7] = line8_new        

        with open(new_path, "w") as f:
            f.writelines(lines)

        #changing txt file back into netlist file
        old_path1 = (file_path + "\\AC_Simulation_Result.txt")
        new_path1 = (file_path + "\\AC_Simulation_Result.net")
        
        dir_path, filename = os.path.split(old_path1)
        name, ext = os.path.splitext(filename)
        new_name, new_ext = os.path.splitext(new_path1)
        new_filename = new_name + new_ext
        new_path = os.path.join(dir_path, new_filename)
        shutil.copy(old_path1, new_path1)

        LTR = SimCommander(file_path + "\\AC_Simulation_Result.net")
        LTR.run()
        LTR.wait_completion()

        # Parse the LTSpice raw file
        t = ltspice.Ltspice(file_path + "\\AC_Simulation_Result_1.raw")
        t.parse()

        # Get the V(out) trace data
        freq = t.get_frequency()
        Vout = t.get_data("V(out)")

        # Create a DataFrame with the frequency and V(onoise) data
        data = {'Frequency (Hz)': freq, 'V': Vout}
        df = pd.DataFrame(data)

        # Export the DataFrame to an Excel file
        ltspice_output_path = (project_path + '\\' + device + '_Output_Impedance.xlsx')
        df.to_excel(ltspice_output_path, index=False, engine='openpyxl')

        #filter data 
        workbook = openpyxl.load_workbook(ltspice_output_path)
        worksheet = workbook['Sheet1']
        worksheet.cell(row=1, column=4, value='V(out)')

        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
            cell_value = str(row[0].value)
            if '+' in cell_value:
                split_values = cell_value.split('+')
                row[0].value = split_values[0][1:]
                row[0].offset(column=1).value = split_values[1].strip()[:-1]
            elif '-' in cell_value:
                split_values = cell_value.split('-')
                row[0].value = split_values[0][1:]
                row[0].offset(column=1).value = "-" + split_values[1].strip()[:-1]

        # Changing format from Cartesian to Polar
        # In this script, I use the sympy library to handle complex number calculations symbolically. The sympy.sympify() function is used to convert the complex number strings 
        # by replacing the "j" with "j*I" to indicate the imaginary unit. The sympy.sqrt() function is then used to calculate the square root, and the .evalf() method is used to 
        # evaluate the result numerically.
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=4):
            b_value = sympy.sympify(re.sub(r'j$', 'j*I', row[0].value)) if row[0].value else 0
            c_value = sympy.sympify(re.sub(r'j$', 'j*I', row[1].value)) if row[1].value else 0
            v_out = sympy.sqrt(abs(b_value)**2 + abs(c_value)**2).evalf()
            row[2].value = str(v_out)       

        workbook.save(ltspice_output_path)

        fd = pd.read_excel(ltspice_output_path, sheet_name='Sheet1') 
        column_name = 'V(out)'
        #d[column_name] = df[column_name].astype(object)
        fd[column_name] = pd.to_numeric(fd[column_name], errors='coerce')  
        fd.to_excel(ltspice_output_path, index=False)

    
        # Converting the Amplifier - Input Referred Noise.csv to .xlsx
        path_file = pd.read_csv(project_path + '\\' + device + ' extracted files' + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Input and Output Impedance.csv')
        nimble_output_path = project_path + '\\' + device + ' Amplifier - Input and Output Impedance.xlsx'
        path_file.to_excel(nimble_output_path, index=None, header=True)

        # Deleting the extra collumns
        file = openpyxl.load_workbook(nimble_output_path)
        sheet_obj = file.active
        sheet_obj.delete_cols(2)
        sheet_obj.delete_cols(2)
        sheet_obj.delete_cols(3)
        file.save(nimble_output_path)

        # Getting the data from Output_Impedance.xlsx to Amplifier - Input and Output Impedance.xlsx in a new sheet
        wb1 = openpyxl.load_workbook(filename=ltspice_output_path)
        ws1 = wb1.worksheets[0]
        wb2 = openpyxl.load_workbook(filename=nimble_output_path)
        ws2 = wb2.create_sheet(ws1.title)

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save(nimble_output_path)

        # Transfering data from the new sheet to the sheet where the graph will be
        xl = openpyxl.load_workbook(nimble_output_path)
        sheet1 = xl['Sheet11']
        sheet2 = xl['Sheet1']

        columnA = []
        for i in range(1, 1003, 1):
            columnA.append(sheet1.cell(row=i, column=1).value)
        for i in range(1, 1003, 1):
            for i in range(1, 1003, 1):
                sheet2.cell(row=i, column=4).value = columnA[i - 1]

        columnB = []
        for i in range(1, 1003, 1):
            columnB.append(sheet1.cell(row=i, column=4).value)
        for i in range(1, 1003, 1):
            for i in range(1, 1003, 1):
                sheet2.cell(row=i, column=5).value = columnB[i - 1]

        if 'Sheet11' in xl.sheetnames:
            xl.remove(xl['Sheet11'])
        xl.save(nimble_output_path)

        #Deletes unnecessary files after extracting
        if os.path.exists(ltspice_output_path):
            os.remove(ltspice_output_path)
        else:
            print("LTspice file does not exist")

        if os.path.exists(project_path + '\\' + device + ' extracted files'):
            shutil.rmtree(project_path + '\\' + device + ' extracted files')
        else:
            print("Extracted files do not exist")

        #Workbook is created
        old_path = project_path + '\\' + device + ' Amplifier - Input and Output Impedance.xlsx'
        results_file = project_path + '\\' + device + '_G' + gain + '_Result.xlsx'
        os.replace(old_path, results_file)

        workbook = load_workbook(results_file)
        sheet = workbook['Sheet1']
        sheet.title = ('G' + gain)

        sheet.cell(row=1, column=1).value = "Nimble - Freq."
        sheet.cell(row=1, column=2).value = "Nimble - Mag."
        sheet.delete_cols(3)
        sheet.cell(row=1, column=3).value = "LTSpice - Freq."
        sheet.cell(row=1, column=4).value = "LTSpice - Mag."
        sheet.delete_cols(5)
        sheet.cell(row=1, column=5).value = "Datasheet Freq."
        sheet.cell(row=1, column=6).value = "Datasheet Mag." 

        for i in range(1, 21):
            sheet.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)

        x_nimble = Reference(sheet, min_col=2, min_row=2, max_row=1002)
        y_nimble = Reference(sheet, min_col=1, min_row=2, max_row=1002)
        x_ltspice = Reference(sheet, min_col=4, min_row=2, max_row=1002)
        y_ltspice = Reference(sheet, min_col=3, min_row=2, max_row=1002)
        x_datasheet = Reference(sheet, min_col=6, min_row=2, max_row=1002)
        y_datasheet = Reference(sheet, min_col=5, min_row=2, max_row=1002)

        series_voltage = Series(x_nimble, y_nimble,title_from_data=False, title="Nimble")
        series_freq = Series(x_ltspice, y_ltspice,title_from_data=False, title="LTspice")
        series_datasheet = Series(x_datasheet, y_datasheet, title_from_data=False, title="Datasheet")

        chart = ScatterChart()
        chart.series.append(series_voltage)
        chart.series.append(series_freq)
        chart.series.append(series_datasheet)

        chart.x_axis.scaling.logBase = 10
        chart.y_axis.scaling.logBase = 10
        chart.y_axis.crossesAt = 0.1
        chart.x_axis.number_format = '0E+00'
        #chart.y_axis.number_format = '0E+00'
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.tickLblSkip = 3

        chart.x_axis.scaling.min = float(self.testData['x_axis_min'])
        chart.y_axis.scaling.min = float(self.testData['y_axis_min'])
        chart.x_axis.scaling.max = float(self.testData['x_axis_max'])
        chart.y_axis.scaling.max = float(self.testData['y_axis_max'])
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.minorGridlines = ChartLines()
        chart.y_axis.minorGridlines = ChartLines()
        chart.height = 12 
        chart.width = 22

        chart.title = "Output Impedance vs. Frequency"
        chart.x_axis.title = self.testData['x_axis_title']
        chart.y_axis.title = self.testData['y_axis_title']
        chart.legend.position = 'r'

        sheet.add_chart(chart, 'I3')
        link = driver.current_url
        sheet['I28'] = link

        for col in range(1, 7):
            cell = sheet.cell(row=2, column=col)
            cell.alignment = Alignment(wrap_text=True)
            sheet.column_dimensions[get_column_letter(col)].width = 15

        workbook.save(results_file)

        print('Graph Complete')

        # Transfering datasheet from the Source file to the Results file
        device_name = device.lower()
        name_string = "_WithScores.xlsx"
        matching_files = [filename for filename in os.listdir(project_path) if device_name in filename.lower() and name_string in filename]

        xl = openpyxl.load_workbook(results_file)
        destination_ws = xl.worksheets[0]

        if matching_files:
            source_file = os.path.join(project_path, matching_files[0])
            xl1 = openpyxl.load_workbook(source_file)
            source_ws = xl1['Datasheet']

            headers = []
            for cell in source_ws[1]:
                headers.append(cell.value)

            columnA_index = headers.index('G' + gain + ' freq')
            columnB_index = headers.index('G' + gain + ' values')

            for i, row in enumerate(source_ws.iter_rows(min_row=2)):
                destination_ws.cell(row=i+2, column=5).value = row[columnA_index].value
                destination_ws.cell(row=i+2, column=6).value = row[columnB_index].value

        else:
            raise Exception(device + " Datasheet source file does not exist")

        print("Copying data from Datasheet")

        gain_sheet = xl['G' + gain]

        #Apply formula to column D - LTSpice from gain_sheet if necessary
        if self.testData['transform_LTSpice_to_dB'] == 'Yes':
            for cell in gain_sheet['D']:
                if isinstance(cell.value, (int, float)):
                    cell.value = 10 ** (cell.value/20)
        else:
            print("No formula applied to LTSpice data")            

        xl.create_sheet('G' + gain + ' Score')
        xl.save(results_file)

        #Customize Scoring sheet
        score_sheet = ('G' + gain + ' Score')
        xl.active = xl[score_sheet]

        cell_ranges = ['A1:D1', 'E1:L1', 'M1:T1']
        texts = ['Info for score', 'Nimble score', 'LTspice score']

        for cell_range, text in zip(cell_ranges, texts):
            xl.active.merge_cells(cell_range)
            cell = xl.active[cell_range.split(':')[0]]
            cell.value = text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        xl.active['A2'] = 'Magnitude range'
        xl.active['B2'] = 'Frequency range'
        xl.active['C2'] = 'Datasheet freq'
        xl.active['D2'] = 'Datasheet mag'

        xl.active['E2'] = 'Closest match without going over index'
        xl.active['F2'] = 'Below freq'
        xl.active['G2'] = 'Above freq'
        xl.active['H2'] = 'Below mag'
        xl.active['I2'] = 'Above mag'
        xl.active['J2'] = 'Linear interpolation'
        xl.active['K2'] = 'Error (dB)'
        xl.active['L2'] = 'Score'
        xl.active['L2'].font = Font(bold=True)

        xl.active['M2'] = 'Closest match without going over index'
        xl.active['N2'] = 'Below freq'
        xl.active['O2'] = 'Above freq'
        xl.active['P2'] = 'Below mag'
        xl.active['Q2'] = 'Above mag'
        xl.active['R2'] = 'Linear interpolation'
        xl.active['S2'] = 'Error (dB)'
        xl.active['T2'] = 'Score'
        xl.active['T2'].font = Font(bold=True)

        xl.active['A3'] = self.testData['y_axis_min']
        xl.active['A4'] = self.testData['y_axis_max']

        xl.active['B3'] = self.testData['x_axis_min']
        xl.active['B4'] = self.testData['x_axis_max']
        
        #Setting columns width
        for col in range(1, 22):
            cell = xl.active.cell(row=2, column=col)
            cell.alignment = Alignment(wrap_text=True)
            xl.active.column_dimensions[get_column_letter(col)].width = 13

        xl.save(results_file)

        sheet_Gx = 'G'+ gain
        sheet_Gx_Score = 'G' + gain + ' Score'

        def copy_ranges_within_excel(workbook, source_sheet, target_sheet, target_col1, target_col2):
            wb = openpyxl.load_workbook(workbook)
            ws_source = wb[source_sheet]
            ws_target = wb[target_sheet]

            max_row = ws_source.max_row

            for row in range(2, max_row + 1):
                # Copy data from column G
                cell_value = ws_source.cell(row=row, column=5).value
                if cell_value is not None:
                    ws_target.cell(row=row + 1, column=target_col1).value = cell_value

                # Copy data from column H
                cell_value = ws_source.cell(row=row, column=6).value
                if cell_value is not None:
                    ws_target.cell(row=row + 1, column=target_col2).value = cell_value

            wb.save(workbook)

        # Call the function with appropriate arguments
        copy_ranges_within_excel(results_file, sheet_Gx, sheet_Gx_Score, 3, 4)

        #This fuction applies the formulas to create the score for Nimble and LTspice
        def apply_formulas(workbook, sheet1_name, sheet2_name):
            wb = openpyxl.load_workbook(workbook)

            sheet1 = wb[sheet1_name]
            sheet2 = wb[sheet2_name]

            max_row = sheet1.max_row
            column_C = 'C'

            # Find the last non-empty row in column C
            for row in range(max_row, 0, -1):
                if sheet1[f"{column_C}{row}"].value is not None:
                    max_row = row
                    break

            #This function will determine valid data in range, on which scoring will be applied
            x_range = [float(self.testData['x_axis_min']), float(self.testData['x_axis_max'])] 
            y_range = [float(self.testData['y_axis_min']), float(self.testData['y_axis_max'])] 

            valid_rows = []
            for row in sheet1.iter_rows(min_row=3):
                if row[2].value is not None and row[3].value is not None and x_range[0] <= row[2].value <= x_range[1] and y_range[0] <= row[3].value <= y_range[1]:
                    valid_rows.append(row[0].row)
            valid_rows_range_l = ["K" + str(num) for num in valid_rows] 
            valid_rows_list_l = ("{0}".format(', '.join(map(str, valid_rows_range_l))))
            valid_rows_range_t = ["S" + str(num) for num in valid_rows] 
            valid_rows_list_t = ("{0}".format(', '.join(map(str, valid_rows_range_t))))

            # Formulas for Nimble scoring
            for row in range(3, max_row+1):
                cell_e = sheet1.cell(row=row, column=5)  # Column 5 corresponds to 'E' =MATCH(C3,'G2'!$A$2:$A$432,1)
                cell_e.value = f'=MATCH(C{row}, INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), 1)'
                cell_f = sheet1.cell(row=row, column=6)  # Column 6 corresponds to 'F' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3)
                cell_f.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), E{row})'             
                cell_g = sheet1.cell(row=row, column=7)  # Column 7 corresponds to 'G' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3+1)
                cell_g.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), E{row}+1)'              
                cell_h = sheet1.cell(row=row, column=8)  # Column 8 corresponds to 'H' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3)
                cell_h.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$B$2:$B$432"), E{row}))'          
                cell_i = sheet1.cell(row=row, column=9)  # Column 9 corresponds to 'I' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3+1)
                cell_i.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$B$2:$B$432"), E{row}+1))'               
                cell_j = sheet1.cell(row=row, column=10)  # Column 10 corresponds to 'J' =SLOPE(H3:I3,F3:G3)*(C3-F3)+H3
                cell_j.value = f'=SLOPE(H{row}:I{row}, F{row}:G{row})*(C{row}-F{row})+H{row}'             
                cell_k = sheet1.cell(row=row, column=11)  # Column 11 corresponds to 'K' =ABS(J3-D3)
                cell_k.value = f'=ABS(J{row}-20*LOG10(D{row}))'
                
            cell_l3 = sheet1.cell(row=3, column=12)  # Column 12 corresponds to 'L' =AVERAGE(K3:K{max_row})
            cell_l3.value = f'=AVERAGE({valid_rows_list_l})'
            cell_l3.font = Font(bold=True)

            # Formulas for LTSpice scoring
            for row in range(3, max_row+1):
                cell_m = sheet1.cell(row=row, column=13)  # Column 5 corresponds to 'M' =MATCH(C3,'G2'!$A$2:$A$432,1)
                cell_m.value = f'=MATCH(C{row}, INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), 1)'
                cell_n = sheet1.cell(row=row, column=14)  # Column 6 corresponds to 'N' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3)
                cell_n.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), M{row})'
                cell_o = sheet1.cell(row=row, column=15)  # Column 7 corresponds to 'O' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3+1)
                cell_o.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), M{row}+1)'
                cell_p = sheet1.cell(row=row, column=16)  # Column 8 corresponds to 'P' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3)
                cell_p.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$D$2:$D$1002"), M{row}))'
                cell_q = sheet1.cell(row=row, column=17)  # Column 9 corresponds to 'Q' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3+1)
                cell_q.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$D$2:$D$1002"), M{row}+1))'
                cell_r = sheet1.cell(row=row, column=18)  # Column 10 corresponds to 'R' =SLOPE(H3:I3,F3:G3)*(C3-F3)+H3
                cell_r.value = f'=SLOPE(P{row}:Q{row}, N{row}:O{row})*(C{row}-N{row})+P{row}'
                cell_s = sheet1.cell(row=row, column=19)  # Column 11 corresponds to 'S' =ABS(J3-D3)
                cell_s.value = f'=ABS(R{row}-20*LOG10(D{row}))'
                
            cell_t3 = sheet1.cell(row=3, column=20)  # Column 12 corresponds to 'T' =AVERAGE(K3:K{max_row})
            cell_t3.value = f'=AVERAGE({valid_rows_list_t})'
            cell_t3.font = Font(bold=True)
                
            wb.save(workbook)

        apply_formulas(results_file, sheet_Gx_Score, sheet_Gx)

        #move result file to results folder
        results_folder = 'Automated_Test_Results'
        r_file = device + '_G' + gain + '_Result.xlsx'

        if not os.path.exists(os.path.join(project_path, results_folder)):
            os.makedirs(os.path.join(project_path, results_folder))

        shutil.move(os.path.join(project_path, r_file), os.path.join(project_path, results_folder, r_file))
        
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        print("                 Scoring Complete!                       ")
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

    def tearDown(self):
        self.driver.quit()

if __name__ == '__main__':
    unittest.main()        
