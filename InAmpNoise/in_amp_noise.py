import filecmp
import math
import zipfile
import json
import time
import os
import shutil
from datetime import datetime
import unittest
import pyautogui
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
import pywinauto.keyboard
import pywinauto.mouse
import pandas as pd
import csv
import openpyxl
import string
from openpyxl import load_workbook
from openpyxl.chart import  (ScatterChart, Reference, Series)
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import range_boundaries
import openpyxl.utils.cell
import ltspice
from PyLTSpice import SimCommander
import numpy as np
import sys

class InAmp(unittest.TestCase):

    def setUp(self):
        # driver instance
        options = Options()
        options.add_argument("--headless=new")
        chromedriver_autoinstaller.install()
        self.driver = webdriver.Chrome(options=options)
        with open(r'inAmpNoise.json') as d:
            self.testData = json.load(d)['Nimble'][0]

    def test_export(self):
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        print("            inAmpNoise script is running...              ")
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        driver = self.driver
        driver.maximize_window()
        driver.get(self.testData['URL'])
        #run the simulation in Nimble

        #Accept Cookies
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#noise-spinner")))
        WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.CSS_SELECTOR, "#noise-spinner")))
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.consent-dialog:nth-child(1) div.modal.fade.in.show "
                             "div.modal-dialog div.modal-content div.modal-body div.short-description > a.btn.btn-success:nth-child(2)"))).click()

        #click on already dragged Amp
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((
            By.XPATH,"//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']"
                     "/div[@id='build-signal-chain-tab-content']/div[@id='adi-signal-chain-row']/div[@id='analog-signal-chain-group']"
                     "/div[@id='signal-chain-drop-area']/table[1]/tr[1]/td[1]/div[1]/div[2]/div[2]/*[1]"))).click()
        
        gain = self.testData['gain']

        # based on the json file the script will select a SingleEnded inAmp or a DifferentialOutput inAmp.
        if (self.testData['amp_type'] == 'Single Ended'):
            print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print("          SingleEnded Output inAmp was selected.         ")
            print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#tspan2988-5"))).click()
            #time.sleep(1)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-0"))).send_keys(self.testData['device'])

            #check if part is present or not disabled in Nimble list
            try:
                element = driver.find_element(By.CSS_SELECTOR, "#device-table > div.slick-pane.slick-pane-top.slick-pane-left > div.slick-viewport.slick-viewport-top.slick-viewport-left > div > div")
                class_attribute = element.get_attribute('class')
                if class_attribute and 'disabled' in class_attribute:
                    raise Exception((self.testData['device']) + " can't be selected in Nimble list")
                else:
                    element.click()
            except NoSuchElementException:
                raise Exception((self.testData['device']) + " can't be selected in Nimble list")
            
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']"
                          "/div[@id='config-signal-chain-item-modal']/div[1]/div[1]/div[1]/div[2]/div[1]/div[4]/div[1]/div[1]/div[1]"
                          "/div[3]/div[1]/button[1]"))).click()
            #select gain value
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input' ))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(Keys.CONTROL + "a")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(Keys.DELETE)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(self.testData['gain'])
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']"
                          "/div[@id='config-signal-chain-item-modal']/div[1]/div[1]/div[1]/div[3]/div[1]/button[1]"))).click()

        elif (self.testData['amp_type'] == 'Differential Output'):
            print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print("          Differential Output inAmp was selected.        ")
            print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.XPATH, "//div[contains(text(),'►')]"))).click()
            #select Differential In Differential Out Configuration
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "body.ember-application.modal-open:nth-child(2) div.adi-modal.modal-fills-window:nth-child(5) " \
                                "div.modal.fade.show.d-block:nth-child(1) div.modal-dialog div.modal-content div.modal-body " \
                                "div.configure-amp.configure-signal-chain-item div.top-area section.config-section div.config-area:nth-child(2) " \
                                "div.svg-button-group.wiring-button-container > div.svg-button-container:nth-child(5)"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "body.ember-application.modal-open:nth-child(2) div.adi-modal.modal-fills-window:nth-child(5) "
                                 "div.modal.fade.show.d-block:nth-child(1) div.modal-dialog div.modal-content div.modal-body "
                                 "div.configure-amp.configure-signal-chain-item div.top-area section.config-section div.config-area:nth-child(2) "
                                 "div.config-more-inputs div.left-right div.amp-sub-type-select:nth-child(1) div.adi-radio:nth-child(3) "
                                 "label:nth-child(1) > input:nth-child(1)"))).click()
            #choose part
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#tspan2988-4-54-5"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-0"))).send_keys(self.testData['device'])

            #check if part is present or not disabled in Nimble list
            try:
                element = driver.find_element(By.CSS_SELECTOR, "#device-table > div.slick-pane.slick-pane-top.slick-pane-left > div.slick-viewport.slick-viewport-top.slick-viewport-left > div > div")
                class_attribute = element.get_attribute('class')
                if class_attribute and 'disabled' in class_attribute:
                    raise Exception((self.testData['device']) + " can't be selected in Nimble list")
                else:
                    element.click()
            except NoSuchElementException:
                raise Exception((self.testData['device']) + " can't be selected in Nimble list")
            
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "body.ember-application.modal-open:nth-child(2) div.adi-modal.modal-fills-window:nth-child(5) div.modal.fade.show.d-block:nth-child(1) "
                                 "div.modal-dialog div.modal-content div.modal-body div.configure-amp.configure-signal-chain-item div.adi-modal.modal-fills-window:nth-child(5) "
                                 "div.modal.fade.show.d-block:nth-child(1) div.modal-dialog div.modal-content div.modal-footer div.button-row > button.btn.btn-primary:nth-child(1)"))).click()
            #choose Input Gain
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-discrete-gain-input"))).click()
            
            gain_dropdown = driver.find_element(By.CSS_SELECTOR, "#amp-discrete-gain-input")
            dropdown = Select(gain_dropdown)
            dropdown.select_by_value(gain)            

            #choose Common Mode Out
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#common-mode-out-input"))).click()
            driver.execute_script(f"document.querySelector('#common-mode-out-input').value={self.testData['common_mode_out']}")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#config-signal-chain-item-modal-title"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "#config-signal-chain-item-modal > div.modal.fade.show.d-block > div > div > div.modal-footer.signal-chain-base-modal-footer > div > button.btn.btn-primary"))).click()
        else:
            raise Exception("Incorrect Amp type selected in JSON file")

        # download All files from Nimble
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/"
                      "div[@id='next-steps-tab-content']/div[@id='next-steps-container']/div[2]/div[1]"))).click()
        time.sleep(5)

        #This script is extracting files directly to project folder
        now = datetime.now()
        day = str(now.day)
        current_date = now.strftime(f"%B {day}, %Y")
        downloads_path = self.testData['downloads_path']
        project_path = self.testData['project_location']
        device = self.testData['device']
        file_path = downloads_path + 'Full Data Export - ' + current_date + '.zip'
        with zipfile.ZipFile(file_path) as zip_ref:
            new_path = project_path + '\\' + device + ' extracted files'
            zip_ref.extractall(new_path)
        print("Files were extracted to project folder")

        #Deletes the zip file after extracting
        if os.path.exists(file_path):
            os.remove(file_path)
        else:
            print("Zip file does not exist")

        # Running the simulation in LTSpice
        file_path = project_path + '\\' + device + ' extracted files' + '\\' + 'Ltspice Schematics'
        LTC = SimCommander(file_path + "\\Noise_Simulation.asc")
        LTC.run()
        LTC.wait_completion()

        # Run the LTSpice raw file
        l = ltspice.Ltspice(file_path + "\\Noise_Simulation_1.raw")
        l.parse()

        # Get the V(onoise) trace data
        freq = l.get_frequency()
        Vonoise = l.get_data('V(onoise)')

        # Create a DataFrame with the frequency and V(onoise) data
        data = {'Frequency (Hz)': freq, 'V(onoise)': Vonoise}
        df = pd.DataFrame(data)

        # Export the DataFrame to an Excel file
        ltspice_output_path = (project_path + '\\' + device + ' Noise_Simulation.xlsx')
        df.to_excel(ltspice_output_path, index=False, engine='openpyxl')

        # Converting the Amplifier - Input Referred Noise.csv to .xlsx
        path_file = pd.read_csv(project_path + '\\' + device + ' extracted files' + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Input Referred Noise.csv')
        nimble_output_path = project_path + '\\' + device + ' Amplifier - Input Referred Noise.xlsx'
        path_file.to_excel(nimble_output_path, index=None, header=True)

        # Deleting the extra collumns
        file = openpyxl.load_workbook(nimble_output_path)
        sheet_obj = file.active
        sheet_obj.delete_cols(3)
        sheet_obj.delete_cols(3)
        sheet_obj.delete_cols(3)
        file.save(nimble_output_path)

        # Getting the data from Noise_Simulation.xlsx to Amplifier - Input Referred Noise.xlsx in a new sheet
        wb1 = openpyxl.load_workbook(filename=ltspice_output_path)
        ws1 = wb1.worksheets[0]
        wb2 = openpyxl.load_workbook(filename=nimble_output_path)
        ws2 = wb2.create_sheet(ws1.title)

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save(nimble_output_path)

        # Transfering data from the new sheet to the sheet where the graph will be
        xl = openpyxl.load_workbook(nimble_output_path)
        sheet1 = xl['Sheet11']
        sheet2 = xl['Sheet1']

        columnA = []
        for i in range(1, 1003, 1):
            columnA.append(sheet1.cell(row=i, column=1).value)
        for i in range(1, 1003, 1):
            for i in range(1, 1003, 1):
                sheet2.cell(row=i, column=4).value = columnA[i - 1]

        columnB = []
        for i in range(1, 1003, 1):
            columnB.append(sheet1.cell(row=i, column=2).value)
        for i in range(1, 1003, 1):
            for i in range(1, 1003, 1):
                sheet2.cell(row=i, column=5).value = columnB[i - 1]

        if 'Sheet11' in xl.sheetnames:
            xl.remove(xl['Sheet11'])
        xl.save(nimble_output_path)

        #Deletes unnecessary files after extracting
        if os.path.exists(ltspice_output_path):
            os.remove(ltspice_output_path)
        else:
            print("LTspice file does not exist")

        if os.path.exists(project_path + '\\' + device + ' extracted files'):
            shutil.rmtree(project_path + '\\' + device + ' extracted files')
        else:
            print("Downloaded files do not exist")

        #Workbook is created
        old_path = project_path + '\\' + device + ' Amplifier - Input Referred Noise.xlsx'
        new_path = project_path + '\\' + device + '_G' + self.testData['gain'] + '_Result.xlsx'
        os.replace(old_path, new_path)
        results_file = project_path + '\\' + device + '_G' + self.testData['gain'] + '_Result.xlsx'
        workbook = load_workbook(results_file)
        sheet = workbook['Sheet1']
        sheet.title = ('G' + self.testData['gain'])

        sheet.cell(row=1, column=1).value = "Nimble - Freq."
        sheet.cell(row=1, column=2).value = "Nimble - Mag."
        sheet.delete_cols(3)
        sheet.cell(row=1, column=3).value = "LTSpice - Freq."
        sheet.cell(row=1, column=4).value = "LTSpice - Mag."
        sheet.delete_cols(5)
        sheet.cell(row=1, column=5).value = "Datasheet Freq."
        sheet.cell(row=1, column=6).value = "Datasheet Noise*" + self.testData['gain']

        for i in range(1, 21):
            sheet.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)

        x_nimble = Reference(sheet, min_col=2, min_row=2, max_row=1002)
        y_nimble = Reference(sheet, min_col=1, min_row=2, max_row=1002)
        x_ltspice = Reference(sheet, min_col=4, min_row=2, max_row=1002)
        y_ltspice = Reference(sheet, min_col=3, min_row=2, max_row=1002)
        x_datasheet = Reference(sheet, min_col=6, min_row=2, max_row=1002)
        y_datasheet = Reference(sheet, min_col=5, min_row=2, max_row=1002)

        series_voltage = Series(x_nimble, y_nimble,title_from_data=False, title="Nimble")
        series_freq = Series(x_ltspice, y_ltspice,title_from_data=False, title="LTspice")
        series_datasheet = Series(x_datasheet, y_datasheet, title_from_data=False, title="Datasheet")

        chart = ScatterChart()
        chart.series.append(series_voltage)
        chart.series.append(series_freq)
        chart.series.append(series_datasheet)

        chart.x_axis.scaling.logBase = 10
        chart.y_axis.scaling.logBase = 10
        chart.y_axis.crossesAt = 0.1
        chart.y_axis.number_format = '0E+00'
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.tickLblSkip = 3

        chart.x_axis.scaling.min = float(self.testData['x_axis_min'])
        chart.y_axis.scaling.min = float(self.testData['y_axis_min'])
        chart.x_axis.scaling.max = float(self.testData['x_axis_max'])
        chart.y_axis.scaling.max = float(self.testData['y_axis_max'])
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.minorGridlines = ChartLines()
        chart.y_axis.minorGridlines = ChartLines()
        chart.height = 12 
        chart.width = 22

        chart.title = None
        chart.x_axis.title = self.testData['x_axis_title']
        chart.y_axis.title = self.testData['y_axis_title']
        chart.legend.position = 'r'

        sheet.add_chart(chart, 'I3')
        link = driver.current_url
        sheet['I28'] = link

        for col in range(1, 7):
            cell = sheet.cell(row=2, column=col)
            cell.alignment = Alignment(wrap_text=True)
            sheet.column_dimensions[get_column_letter(col)].width = 15

        workbook.save(results_file)

        print('Graph Complete')

        # Transfering datasheet from the Source file to the Results file
        xl = openpyxl.load_workbook(results_file)
        destination_ws = xl.worksheets[0]

        if os.path.exists(project_path + '\\' + device + '_WithScores.xlsx'):
            xl1 = openpyxl.load_workbook(project_path + '\\' + device + '_WithScores.xlsx')
            source_ws = xl1['Datasheet']

            headers = []
            for cell in source_ws[1]:
                headers.append(cell.value)

            columnA_index = headers.index('G' + self.testData['gain'] + ' freq')
            columnB_index = headers.index('G' + self.testData['gain'] + ' mag')

            for i, row in enumerate(source_ws.iter_rows(min_row=2)):
                destination_ws.cell(row=i+2, column=5).value = row[columnA_index].value
                destination_ws.cell(row=i+2, column=6).value = row[columnB_index].value

        else:
            raise Exception(self.testData['device'] + " Datasheet source file does not exist")

        print("Copying data from Datasheet")

        gain_sheet = xl['G' + self.testData['gain']]

        # Multiply column B - Nimble from gain_sheet by 1e9 and by gain value
        for cell in gain_sheet['B']:
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value * 1e9 * float(self.testData['gain'])

        # Multiply column D - LTspice from gain_sheet by 1e9
        for cell in gain_sheet['D']:
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value * 1e9

        # Multiply column F - Datasheet by gain value
        for cell in gain_sheet['F']:
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value * float(self.testData['gain'])

        xl.create_sheet('G' + self.testData['gain'] + ' Score')
        xl.save(results_file)

        #Customize Scoring sheet
        score_sheet = ('G' + self.testData['gain'] + ' Score')
        xl.active = xl[score_sheet]

        cell_ranges = ['A1:D1', 'E1:L1', 'M1:T1']
        texts = ['Info for score', 'Nimble score', 'LTspice score']

        for cell_range, text in zip(cell_ranges, texts):
            xl.active.merge_cells(cell_range)
            cell = xl.active[cell_range.split(':')[0]]
            cell.value = text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        xl.active['A2'] = 'Magnitude range'
        xl.active['B2'] = 'Frequency range'
        xl.active['C2'] = 'Datasheet freq'
        xl.active['D2'] = 'Datasheet mag'

        xl.active['E2'] = 'Closest match without going over index'
        xl.active['F2'] = 'Below freq'
        xl.active['G2'] = 'Above freq'
        xl.active['H2'] = 'Below mag'
        xl.active['I2'] = 'Above mag'
        xl.active['J2'] = 'Linear interpolation'
        xl.active['K2'] = 'Error (dB)'
        xl.active['L2'] = 'Score'
        xl.active['L2'].font = Font(bold=True)

        xl.active['M2'] = 'Closest match without going over index'
        xl.active['N2'] = 'Below freq'
        xl.active['O2'] = 'Above freq'
        xl.active['P2'] = 'Below mag'
        xl.active['Q2'] = 'Above mag'
        xl.active['R2'] = 'Linear interpolation'
        xl.active['S2'] = 'Error (dB)'
        xl.active['T2'] = 'Score'
        xl.active['T2'].font = Font(bold=True)

        xl.active['A3'] = self.testData['y_axis_min']
        xl.active['A4'] = self.testData['y_axis_max']

        xl.active['B3'] = self.testData['x_axis_min']
        xl.active['B4'] = self.testData['x_axis_max']

        #Setting columns width
        for col in range(1, 22):
            cell = xl.active.cell(row=2, column=col)
            cell.alignment = Alignment(wrap_text=True)
            xl.active.column_dimensions[get_column_letter(col)].width = 13

        xl.save(results_file)

        gain = self.testData['gain']
        sheet_Gx = 'G'+gain
        sheet_Gx_Score = 'G' + gain + ' Score' 

        def copy_ranges_within_excel(workbook, source_sheet, target_sheet, target_col1, target_col2):
            wb = openpyxl.load_workbook(workbook)
            ws_source = wb[source_sheet]
            ws_target = wb[target_sheet]

            max_row = ws_source.max_row

            for row in range(2, max_row + 1):
                # Copy data from column G
                cell_value = ws_source.cell(row=row, column=5).value
                if cell_value is not None:
                    ws_target.cell(row=row + 1, column=target_col1).value = cell_value

                # Copy data from column H
                cell_value = ws_source.cell(row=row, column=6).value
                if cell_value is not None:
                    ws_target.cell(row=row + 1, column=target_col2).value = cell_value

            wb.save(workbook)

        # Call the function with appropriate arguments
        copy_ranges_within_excel(results_file, sheet_Gx, sheet_Gx_Score, 3, 4)

        #This fuction applies the formulas to create the score for Nimble and LTspice
        def apply_formulas(workbook, sheet1_name, sheet2_name):
            wb = openpyxl.load_workbook(workbook)

            sheet1 = wb[sheet1_name]
            sheet2 = wb[sheet2_name]

            max_row = sheet1.max_row
            column_C = 'C'

            # Find the last non-empty row in column C
            for row in range(max_row, 0, -1):
                if sheet1[f"{column_C}{row}"].value is not None:
                    max_row = row
                    break

            #This function will determine valid data in range, on which scoring will be applied
            x_range = [float(self.testData['x_axis_min']), float(self.testData['x_axis_max'])] 
            y_range = [float(self.testData['y_axis_min']), float(self.testData['y_axis_max'])] 

            valid_rows = []
            for row in sheet1.iter_rows(min_row=3):
                if row[2].value is not None and row[3].value is not None and x_range[0] <= row[2].value <= x_range[1] and y_range[0] <= row[3].value <= y_range[1]:
                    valid_rows.append(row[0].row)
            valid_rows_range_l = ["K" + str(num) for num in valid_rows] 
            valid_rows_list_l = ("{0}".format(', '.join(map(str, valid_rows_range_l))))
            valid_rows_range_t = ["S" + str(num) for num in valid_rows] 
            valid_rows_list_t = ("{0}".format(', '.join(map(str, valid_rows_range_t))))

            # Formulas for Nimble scoring
            for row in range(3, max_row+1):
                cell_e = sheet1.cell(row=row, column=5)  # Column 5 corresponds to 'E' =MATCH(C3,'G2'!$A$2:$A$432,1)
                cell_e.value = f'=MATCH(C{row}, INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), 1)'
                cell_f = sheet1.cell(row=row, column=6)  # Column 6 corresponds to 'F' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3)
                cell_f.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), E{row})'             
                cell_g = sheet1.cell(row=row, column=7)  # Column 7 corresponds to 'G' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3+1)
                cell_g.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), E{row}+1)'              
                cell_h = sheet1.cell(row=row, column=8)  # Column 8 corresponds to 'H' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3)
                cell_h.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$B$2:$B$432"), E{row}))'          
                cell_i = sheet1.cell(row=row, column=9)  # Column 9 corresponds to 'I' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3+1)
                cell_i.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$B$2:$B$432"), E{row}+1))'               
                cell_j = sheet1.cell(row=row, column=10)  # Column 10 corresponds to 'J' =SLOPE(H3:I3,F3:G3)*(C3-F3)+H3
                cell_j.value = f'=SLOPE(H{row}:I{row}, F{row}:G{row})*(C{row}-F{row})+H{row}'             
                cell_k = sheet1.cell(row=row, column=11)  # Column 11 corresponds to 'K' =ABS(J3-D3)
                cell_k.value = f'=ABS(J{row}-20*LOG10(D{row}))'
                
            cell_l3 = sheet1.cell(row=3, column=12)  # Column 12 corresponds to 'L' =AVERAGE(K3:K{max_row})
            cell_l3.value = f'=AVERAGE({valid_rows_list_l})'
            cell_l3.font = Font(bold=True)
            
            # Formulas for LTSpice scoring
            for row in range(3, max_row+1):
                cell_m = sheet1.cell(row=row, column=13)  # Column 5 corresponds to 'M' =MATCH(C3,'G2'!$A$2:$A$432,1)
                cell_m.value = f'=MATCH(C{row}, INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), 1)'
                cell_n = sheet1.cell(row=row, column=14)  # Column 6 corresponds to 'N' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3)
                cell_n.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), M{row})'
                cell_o = sheet1.cell(row=row, column=15)  # Column 7 corresponds to 'O' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3+1)
                cell_o.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), M{row}+1)'
                cell_p = sheet1.cell(row=row, column=16)  # Column 8 corresponds to 'P' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3)
                cell_p.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$D$2:$D$1002"), M{row}))'
                cell_q = sheet1.cell(row=row, column=17)  # Column 9 corresponds to 'Q' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3+1)
                cell_q.value = f'=20*LOG10(INDEX(INDIRECT("\'{sheet2_name}\'!$D$2:$D$1002"), M{row}+1))'
                cell_r = sheet1.cell(row=row, column=18)  # Column 10 corresponds to 'R' =SLOPE(H3:I3,F3:G3)*(C3-F3)+H3
                cell_r.value = f'=SLOPE(P{row}:Q{row}, N{row}:O{row})*(C{row}-N{row})+P{row}'
                cell_s = sheet1.cell(row=row, column=19)  # Column 11 corresponds to 'S' =ABS(J3-D3)
                cell_s.value = f'=ABS(R{row}-20*LOG10(D{row}))'
                
            cell_t3 = sheet1.cell(row=3, column=20)  # Column 12 corresponds to 'T' =AVERAGE(K3:K{max_row})
            cell_t3.value = f'=AVERAGE({valid_rows_list_t})'
            cell_t3.font = Font(bold=True)
                
            wb.save(workbook)

        apply_formulas(results_file, sheet_Gx_Score, sheet_Gx)

        #move result file to results folder
        results_folder = 'Automated_Test_Results'
        r_file = device + '_G' + self.testData['gain'] + '_Result.xlsx'

        if not os.path.exists(os.path.join(project_path, results_folder)):
            os.makedirs(os.path.join(project_path, results_folder))

        shutil.move(os.path.join(project_path, r_file), os.path.join(project_path, results_folder, r_file))
        
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        print("                 Scoring Complete!                       ")
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

    def tearDown(self):
        self.driver.quit()

if __name__ == '__main__':
    unittest.main()