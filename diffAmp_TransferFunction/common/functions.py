from datetime import date
import math
import os
import openpyxl

class functions():  

    def get_variables_from_excel(self, file_path):       
        workbook = openpyxl.load_workbook(file_path)
        # Select the worksheet (by default, the first worksheet will be selected)
        worksheet = workbook.active
        dictionaries = []
        for col in worksheet.iter_cols(min_row=2, max_row=2):                
            cell_value = col[0].value                
            if cell_value:
                # Split the cell value using a delimiter (e.g., comma)
                pairs = cell_value.split(',')
                # Store the variables and their values in a dictionary
                variables = {}
                for pair in pairs:
                    variable_name, value = pair.strip().split(':')
                    variables[variable_name] = value
                dictionaries.append(variables)
        return dictionaries
    
    def copy_columns_between_excels(self,
        excel_file_1, excel_file_2,
        sheet_1, column1_1, column2_1,
        sheet_2, column1_2, column2_2):
        # Load the workbooks and worksheets
        wb1 = openpyxl.load_workbook(excel_file_1)
        ws1 = wb1[sheet_1]
        wb2 = openpyxl.load_workbook(excel_file_2)
        ws2 = wb2[sheet_2]
        # Copy the specified columns from sheet_1 to sheet_2
        for row in range(1, ws1.max_row + 1):
            ws2.cell(row=row, column=column1_2).value = ws1.cell(row=row, column=column1_1).value
            ws2.cell(row=row, column=column2_2).value = ws1.cell(row=row, column=column2_1).value
        # Save the changes to excel_file_2
        wb2.save(excel_file_2)
    
    def get_formatted_current_date(self):
        today = date.today()
        day = str(today.day)
        current_date = today.strftime(f"%B {day}, %Y")
        return current_date
    
    def value_to_position(self, value, limitmin, limitmax):
        minpos = 1
        maxpos = 10000
        minval = math.log(limitmin)
        maxval = math.log(limitmax)
        scale = (maxval - minval) / (maxpos - minpos)
        if value <= 0:
            return minpos
        else:
            position = minpos + (math.log(value) - minval) / scale
            return position
        
    def text_to_num(self, si_string):
        si_prefixes = {
            'y': 1e-24,  # yocto
            'z': 1e-21,  # zepto
            'a': 1e-18,  # atto
            'f': 1e-15,  # femto
            'p': 1e-12,  # pico
            'n': 1e-9,   # nano
            'u': 1e-6,   # micro
            'm': 1e-3,   # milli
            'k': 1e3,    # kilo
            'M': 1e6,    # mega
        }        
        if si_string[-1] in si_prefixes:
            value = float(si_string[:-1]) * si_prefixes[si_string[-1]]
        else:
            value = float(si_string)
        return value 
    
    def create_excel_file(self, folder_path, file_name):
        workbook = openpyxl.Workbook()
        file_path = os.path.join(folder_path, file_name)
        workbook.save(file_path)

    #This functions moves data from a sheet to another in the same excel file            
    def copy_ranges_within_excel(workbook, source_sheet, target_sheet, source_col1, source_col2, target_col1, target_col2, offset_source_sheet, offset_target_sheet):
        wb = openpyxl.load_workbook(workbook)
        ws_source = wb[source_sheet]
        ws_target = wb[target_sheet]
        max_row = ws_source.max_row

        for row in range(1 + offset_source_sheet, max_row + 1):
            # Copy data from source_col1
            cell_value = ws_source.cell(row=row, column=source_col1).value
            if cell_value is not None:
                ws_target.cell(row=row+offset_target_sheet-offset_source_sheet, column=target_col1).value = cell_value

            # Copy data from source_col2
            cell_value = ws_source.cell(row=row, column=source_col2).value
            if cell_value is not None:
                ws_target.cell(row=row+offset_target_sheet-offset_source_sheet, column=target_col2).value = cell_value
        wb.save(workbook)

        # # Call the function with appropriate arguments
        # copy_ranges_within_excel(workbook_path, 'Datasheet', gain_sheet_score, 5, 6, 3, 4, offset_source_sheet=1, offset_target_sheet=2)
    