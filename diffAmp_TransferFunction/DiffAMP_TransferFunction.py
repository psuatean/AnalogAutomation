import unittest
import os
import shutil
import zipfile
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from decimal import Decimal
from datetime import date
import math
import time
import json
import chromedriver_autoinstaller
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.chart.axis import ChartLines
import ltspice
from PyLTSpice import SimCommander
import numpy as np
import re
import sympy
from common.functions import functions

class diffAMP(unittest.TestCase):

    def setUp(self):
        # driver instance
        chromedriver_autoinstaller.install()
        options = Options()
        options.add_argument("--headless=new")
        self.driver = webdriver.Chrome(options=options) 
        with open(r'diffAmp_TransferFunction\DiffAmp_TransferFunction.json') as d:
            self.testData = json.load(d)['Nimble'][0]
        
    def test_export(self):
        my_functions = functions() 
        device = self.testData['device']
        project_path = self.testData['project_location']
        source_workbook = (project_path + '\\' + device + '_WithScores.xlsx') 
        dictionaries = my_functions.get_variables_from_excel(source_workbook)
        results_folder = project_path + '\\' + 'Automated_Test_Results'
        results_file = results_folder + '\\' + device + "_Test_Results.xlsx"

        # Create results folder and results file
        if not os.path.exists(project_path + '\\' + 'Automated_Test_Results'):
            os.makedirs(project_path + '\\' + 'Automated_Test_Results')
        my_functions.create_excel_file(results_folder, results_file)

        for dictionary in dictionaries:
        
            driver = self.driver
            driver.maximize_window()
            driver.get(self.testData['URL'])

            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("       DiffAmp_TransferFunction script is running        ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++") 

            device = self.testData['device']
            gain = dictionary['gain']
            R1 = dictionary['R1']
            C1 = dictionary['C1']
            R3 = dictionary['Load']
            filter_freq = dictionary['Filter_freq']
            downloads_path = self.testData['downloads_path']
            project_path = self.testData['project_location']
            
            source_workbook = project_path + '\\' + device + '_WithScores.xlsx'
             
            new_rvalue = my_functions.text_to_num(R1)
            new_c1_value = my_functions.text_to_num(C1) 
            # new_rc1value = my_functions.text_to_num(paths['rc1_value'])
            new_r3value = my_functions.text_to_num(R3) 
              
            # cookies accept
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.consent-dialog:nth-child(1) div.modal.fade.in.show "
                                "div.modal-dialog div.modal-content div.modal-body div.short-description > a.btn.btn-success:nth-child(2)"))).click()
            
            # amplifier settings
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((
                By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/div[@id='build-signal-chain-tab-content']"
                "/div[@id='adi-signal-chain-row']/div[@id='analog-signal-chain-group']/div[@id='signal-chain-drop-area']/table[1]/tr[1]/td[1]/div[1]/div[2]/div[2]/*[1]"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(Keys.CONTROL + "a")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(Keys.DELETE)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(gain)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(Keys.ENTER)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#tspan2988-4-54-5"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-0"))).send_keys(device)

            #check if part is present or not disabled in Nimble list
            try:
                element = driver.find_element(By.CSS_SELECTOR, "#device-table > div.slick-pane.slick-pane-top.slick-pane-left > div.slick-viewport.slick-viewport-top.slick-viewport-left > div > div")
                class_attribute = element.get_attribute('class')
                if class_attribute and 'disabled' in class_attribute:
                    raise Exception(device + " can't be selected in Nimble list")
                else:
                    element.click()
            except NoSuchElementException:
                raise Exception(device + " can't be selected in Nimble list")
            
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, 'body.ember-application.modal-open:nth-child(2) div.adi-modal.modal-fills-window.modal-hide-scroll:nth-child(5) div.modal.fade.show.d-block:nth-child(1) '
                'div.modal-dialog div.modal-content div.modal-body div.configure-amp.configure-signal-chain-item div.adi-modal.modal-fills-window:nth-child(5) '
                'div.modal.fade.show.d-block:nth-child(1) div.modal-dialog div.modal-content div.modal-footer div.button-row > button.btn.btn-primary:nth-child(1)' ))).click()         
                     
            rposition = my_functions.value_to_position(new_rvalue, 1e1, 1e7)
            c1position = my_functions.value_to_position(new_c1_value, 1e-13, 1e-6)            
            # rc1position = value_to_position(new_rc1value, 150, 1e7)
            r3position = my_functions.value_to_position(new_r3value, 150, 1e7)

            driver.execute_script(f"document.querySelector('#rscale-slider').value = {rposition}; document.querySelector('#rscale-slider').dispatchEvent(new Event('input'));")
            driver.execute_script(f"document.querySelector('#c1-slider').value = {c1position}; document.querySelector('#c1-slider').dispatchEvent(new Event('input'));")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "#config-signal-chain-item-modal > div.modal.fade.show.d-block > div > div > div.modal-footer.signal-chain-base-modal-footer > div > button.btn.btn-primary"))).click()
            
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("                    Slider values set!                   ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")        

            #Set Up Filter values 
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#signal-chain-drop-area #circuit-content[title=\"Filter\"]"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-inputs-type-tab-button"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-type-category-radio-group > div:nth-child(3) > label > input[type=radio]"))).click()        
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#hp-diff-wiring-button"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-order-radio-group > div:nth-child(1) > label > input[type=radio]"))).click()        
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#fp-input"))).send_keys(Keys.CONTROL + "a")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#fp-input"))).send_keys(Keys.DELETE)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#fp-input"))).send_keys(filter_freq)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "#config-signal-chain-item-modal > div.modal.fade.show.d-block > div > div > div.modal-body.signal-chain-base-modal-body > div > div.top-area"
                " > section.config-section > div > div.sub-tab-content-container > button.tab-button-area.enabled.next > div > svg"))).click()

            #driver.execute_script(f"document.querySelector('#rc-r1-slider').value = {rc1position}; document.querySelector('#rc-r1-slider').dispatchEvent(new Event('input'));")
            driver.execute_script(f"document.querySelector('#rc-r3-slider').value = {r3position}; document.querySelector('#rc-r3-slider').dispatchEvent(new Event('input'));")        
            
            driver.execute_script("document.querySelector('#config-signal-chain-item-modal > div.modal.fade.show.d-block > div > div > div.modal-footer > div > button.btn.btn-primary').click()")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
            
            #This script is moving the downloaded files to the project folder
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-container > div.download-area > div.download-all-button > div > h5"))).click()
            time.sleep(5)         

            today = date.today()
            day = str(today.day)
            current_date = today.strftime(f"%B {day}, %Y")
            zip_file_path = downloads_path + '\\' + 'Full Data Export - ' + current_date + '.zip'
            with zipfile.ZipFile(zip_file_path) as zip_ref:
                new_path = project_path + '\\' + device + ' extracted files'
                zip_ref.extractall(new_path)
            print("Files were extracted to project folder")

            #Deletes the zip file after extracting
            if os.path.exists(zip_file_path):
                os.remove(zip_file_path)
            else:
                print("Zip file does not exist")   

            # Move Transfer Function csv to extracted files folder
            raw_data = project_path + '\\' + device + ' extracted files' + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Transfer Function.csv'
            shutil.copy2(raw_data, new_path)

            # Converting the Transfer Function .csv to .xlsx
            path_file = pd.read_csv(project_path + '\\' + device + ' extracted files' + '\\' + 'Amplifier - Transfer Function.csv')
            path_file.to_excel(project_path + '\\' + device + ' extracted files' + '\\' + 'Amplifier - Transfer Function.xlsx', index=None, header=True)
            nimble_output_path = project_path + '\\' + device + ' extracted files' + '\\' + 'Amplifier - Transfer Function.xlsx'

            #Running the simulation in LTSpice
            file_path = project_path + '\\' + device + ' extracted files' + '\\' + 'Ltspice Schematics'
            LTC = SimCommander(file_path + "\\AC_Simulation.asc")
            LTC.run()
            LTC.wait_completion()
            
            # parsing LTSpice files
            l = ltspice.Ltspice(file_path + "\\AC_Simulation_1.raw")
            l.parse()

            # Get the V(out) trace data
            freq = l.get_frequency()
            Vout = l.get_data('V(out)')

            # Change from Carthesian to Polar format 
            Vout_dB = 20 * np.log10(np.abs(Vout))

            # Create a DataFrame with the frequency and magnitude (dB) data
            data = {'Frequency (Hz)': freq, 'Magnitude (dB)': Vout_dB}
            df = pd.DataFrame(data)

            # Export the DataFrame to an Excel file
            ltspice_output_path = (new_path + "\\AC_Simulation.xlsx")
            df.to_excel(ltspice_output_path, index=False, engine='openpyxl')    

            # Name the sheet
            file = openpyxl.load_workbook(results_file)
            sheet = file.active
            result_sheet = ("G" + gain + 'RL' + R3 + 'Ω' )
            sheet.title = result_sheet
            #sheet.delete_cols(3)
            file.save(results_file)

            # Getting the Nimble data from the Transfer_Function.xlsx to Result file
            my_functions.copy_columns_between_excels(
                nimble_output_path, results_file,
                'Sheet1', 1, 2,
                result_sheet, 1, 2) 

            # Getting the LTSpice data from the AC_Simulation.xlsx to Result file
            my_functions.copy_columns_between_excels(
                ltspice_output_path, results_file,
                'Sheet1', 1, 2,
                result_sheet, 4, 5)  
            
            # Getting the Datasheet data to Result File
            #source_workbook = (project_path + '\\' + device + '_WithScores.xlsx')
            my_functions.copy_columns_between_excels(
                source_workbook, results_file,
                'Datasheet', 1, 2,
                result_sheet, 7, 8)
            
            # # Copy result sheet to result file
            # my_functions.copy_columns_between_excels(
            #     excel_path, results_file,
            #     result_sheet, result_sheet)
        

        #Creating Scatter graph    
            workbook_path = (project_path + '\\' + device + '\\' + 'Amplifier - Transfer Function.xlsx')
            workbook = load_workbook(workbook_path)
            sheet = workbook['G'+gain]
            link = driver.current_url
            sheet['J1'] = link

            sheet.cell(row=1, column=1).value = "Nimble-Freq."
            sheet.cell(row=1, column=2).value = "Nimble-Mag."
            sheet.cell(row=1, column=4).value = "LTSpice-Freq."
            sheet.cell(row=1, column=5).value = "LTSspice-Mag."
            sheet.cell(row=1, column=7).value = "Datasheet-Freq."
            sheet.cell(row=1, column=8).value = "Datasheet-Mag."

            for i in range(1,21):
                sheet.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)

            x_nimble = Reference(sheet, min_col=2, min_row=2, max_row=1010)
            y_nimble = Reference(sheet, min_col=1, min_row=2, max_row=1010)
            x_ltspice = Reference(sheet, min_col=5, min_row=2, max_row=1010)
            y_ltspice = Reference(sheet, min_col=4, min_row=2, max_row=1010)
            x_datasheet = Reference(sheet, min_col=8, min_row=2, max_row=1010)
            y_datasheet = Reference(sheet, min_col=7, min_row=2, max_row=1010)

            series_nimble = Series(x_nimble, y_nimble, title_from_data=False, title="Nimble")
            series_ltspice = Series(x_ltspice, y_ltspice, title_from_data=False, title="LTspice")
            series_datasheet = Series(x_datasheet, y_datasheet, title_from_data=False, title="Datasheet")
            
            # Chart type
            chart = ScatterChart()
            chart.series.append(series_nimble)
            chart.series.append(series_ltspice)
            chart.series.append(series_datasheet)

            chart.x_axis.scaling.logBase = 10
            chart.y_axis.number_format = '0.00E+00'
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.tickLblSkip = 3

            chart.x_axis.scaling.min = paths['x_axis_min']
            chart.y_axis.scaling.min = paths['y_axis_min']
            chart.x_axis.scaling.max = paths['x_axis_max']
            chart.y_axis.scaling.max = paths['y_axis_max']
            chart.x_axis.tickLblPos = "low"

            chart.title = None
            chart.x_axis.title = 'Frequency (Hz)'
            chart.y_axis.title = 'Magnitude (dB)'
            chart.legend.position = 'r'

            sheet.add_chart(chart, 'J2')
            workbook.save(workbook_path)
            
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("            Scatter Plot chart was created!              ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            
        # Create and Customize Scoring sheet
            workbook.create_sheet('G' + gain + ' Score')
            score_sheet = ('G' + gain + ' Score')
            sheet = workbook[score_sheet]

            # Creating Header
            cell_ranges = ['A1:D1', 'E1:L1', 'M1:T1']
            texts = ['Info for score', 'Nimble score', 'LTspice score']

            for cell_range, text in zip(cell_ranges, texts):
                sheet.merge_cells(cell_range)
                cell = sheet[cell_range.split(':')[0]]
                cell.value = text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Info for score table
            sheet['A2'] = 'Magnitude range'
            sheet['A3'] = float(paths['y_axis_min'])
            sheet['A4'] = float(paths['y_axis_max'])
            sheet['B2'] = 'Frequency range'
            sheet['B3'] = float(paths['x_axis_min'])
            sheet['B4'] = float(paths['x_axis_max'])
            sheet['C2'] = 'Datasheet freq'
            sheet['D2'] = 'Datasheet mag'
            # Nimble Score table
            sheet['E2'] = 'Closest match without going over index'
            sheet['F2'] = 'Below freq'
            sheet['G2'] = 'Above freq'
            sheet['H2'] = 'Below mag'
            sheet['I2'] = 'Above mag'
            sheet['J2'] = 'Linear interpolation'
            sheet['K2'] = 'Error (dB)'
            sheet['L2'] = 'Score'
            # LTspice Score table
            sheet['M2'] = 'Closest match without going over index'
            sheet['N2'] = 'Below freq'
            sheet['O2'] = 'Above freq'
            sheet['P2'] = 'Below mag'
            sheet['Q2'] = 'Above mag'
            sheet['R2'] = 'Linear interpolation'
            sheet['S2'] = 'Error (dB)'
            sheet['T2'] = 'Score'
            sheet['L2'].font = Font(bold=True)
            sheet['T2'].font = Font(bold=True)

            #Wrap cells and set width
            for col in range(1, 22):
                cell = sheet.cell(row=2, column=col)
                cell.alignment = Alignment(wrap_text=True)
                sheet.column_dimensions[get_column_letter(col)].width = 12
            
            workbook.save(workbook_path)
            
        #Transfering Datasheet from G2 to G2 Score
            def copy_ranges_within_excel(workbook_path, source_sheet, source_range1, source_range2, target_sheet, target_range1, target_range2):
                wb = openpyxl.load_workbook(workbook_path)
                ws_source = wb[source_sheet]
                ws_target = wb[target_sheet]
                for (source_range, target_range) in zip([source_range1, source_range2], [target_range1, target_range2]):
                    for row in ws_source[source_range]:
                        for cell in row:
                            target_cell = ws_target.cell(row=cell.row + 1, column=cell.column - 4)
                            target_cell.value = cell.value
                wb.save(workbook_path)

            sheet_Gx = 'G'+gain
            sheet_Gx_Score = 'G' + gain + ' Score'
            copy_ranges_within_excel(workbook_path, sheet_Gx, 'G2:G56', 'H2:H56', sheet_Gx_Score, 'C3:C57', 'D3:D57')

        #This fuction applies the formulas to create the score for Nimble and LTspice
            def apply_formulas(workbook, sheet1_name, sheet2_name):
                wb = openpyxl.load_workbook(workbook)

                sheet1 = wb[sheet1_name]
                sheet2 = wb[sheet2_name]

                # Iterate through cells E3:E56 and F3:F56 in sheet1 and apply the formulas
                for row in range(3, 57):
                    cell_e = sheet1.cell(row=row, column=5)  # Column 5 corresponds to 'E' =MATCH(C3,'G2'!$A$2:$A$432,1)
                    cell_e.value = f'=MATCH(C{row}, {sheet2_name}!$A$2:$A$432, 1)'
                    cell_f = sheet1.cell(row=row, column=6)  # Column 6 corresponds to 'F' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3)
                    cell_f.value = f'=INDEX({sheet2_name}!$A$2:$A$432, E{row})'             
                    cell_g = sheet1.cell(row=row, column=7)  # Column 7 corresponds to 'G' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3+1)
                    cell_g.value = f'=INDEX({sheet2_name}!$A$2:$A$432, E{row}+1)'              
                    cell_h = sheet1.cell(row=row, column=8)  # Column 8 corresponds to 'H' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3)
                    cell_h.value = f'=INDEX({sheet2_name}!$B$2:$B$432, E{row})'               
                    cell_i = sheet1.cell(row=row, column=9)  # Column 9 corresponds to 'I' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3+1)
                    cell_i.value = f'=INDEX({sheet2_name}!$B$2:$B$432, E{row}+1)'               
                    cell_j = sheet1.cell(row=row, column=10)  # Column 10 corresponds to 'J' =SLOPE(H3:I3,F3:G3)*(C3-F3)+H3
                    cell_j.value = f'=SLOPE(H{row}:I{row}, F{row}:G{row})*(C{row}-F{row})+H{row}'             
                    cell_k = sheet1.cell(row=row, column=11)  # Column 11 corresponds to 'K' =ABS(J3-D3)
                    cell_k.value = f'=ABS(J{row}-D{row})'
                    
                cell_l3 = sheet1.cell(row=3, column=12)  # Column 12 corresponds to 'L' =AVERAGE(K3:K46)
                cell_l3.value = '=AVERAGE(K3:K52)'
                cell_l3.font = Font(bold=True)
                
                # Formulas for LTSpice scoring
                for row in range(3, 57):
                    cell_m = sheet1.cell(row=row, column=13)  # Column 5 corresponds to 'M' =MATCH(C3,'G2'!$A$2:$A$432,1)
                    cell_m.value = f'=MATCH(C{row}, {sheet2_name}!$D$2:$D$1002, 1)'
                    cell_n = sheet1.cell(row=row, column=14)  # Column 6 corresponds to 'N' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3)
                    cell_n.value = f'=INDEX({sheet2_name}!$D$2:$D$1002, M{row})'
                    cell_o = sheet1.cell(row=row, column=15)  # Column 7 corresponds to 'O' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3+1)
                    cell_o.value = f'=INDEX({sheet2_name}!$D$2:$D$1002, M{row}+1)'
                    cell_p = sheet1.cell(row=row, column=16)  # Column 8 corresponds to 'P' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3)
                    cell_p.value = f'=INDEX({sheet2_name}!$E$2:$E$1002, M{row})'
                    cell_q = sheet1.cell(row=row, column=17)  # Column 9 corresponds to 'Q' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3+1)
                    cell_q.value = f'=INDEX({sheet2_name}!$E$2:$E$1002, M{row}+1)'
                    cell_r = sheet1.cell(row=row, column=18)  # Column 10 corresponds to 'R' =SLOPE(H3:I3,F3:G3)*(C3-F3)+H3
                    cell_r.value = f'=SLOPE(P{row}:Q{row}, N{row}:O{row})*(C{row}-N{row})+P{row}'
                    cell_s = sheet1.cell(row=row, column=19)  # Column 11 corresponds to 'S' =ABS(J3-D3)
                    cell_s.value = f'=ABS(R{row}-D{row})'
                    
                cell_t3 = sheet1.cell(row=3, column=20)  # Column 12 corresponds to 'T' =AVERAGE(K3:K46)
                cell_t3.value = '=AVERAGE(S3:S52)'
                cell_t3.font = Font(bold=True)
                    
                wb.save(workbook)

            apply_formulas(workbook_path, sheet_Gx_Score, sheet_Gx)     
                
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("               Scoring sheet was created!                ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

            rename_noise_excel = project_path + '\\' + device + '\\' + 'Amplifier - Transfer Function.xlsx'
            new_noise_excel = project_path + '\\' + device + '.xlsx'
            os.rename(rename_noise_excel, new_noise_excel)

            #function that deletes unwated files
            def delete_extra_files(folder: str, goodfile: str, goodfile2: str) -> None:
                for entry in os.listdir(folder):
                    entry_path = os.path.join(folder, entry)
                    if os.path.isfile(entry_path):
                        if entry != goodfile and entry != goodfile2:
                            os.remove(entry_path)
                        else:
                            continue
                    elif os.path.isdir(entry_path):
                        shutil.rmtree(entry_path)

            folder_path = project_path + '\\' + device
            goodfile1 = f"{device}.xlsx"
            goodfile2 = f"{device}_WithScores.xlsx"
            print(goodfile1, goodfile2)
            delete_extra_files(folder_path, goodfile1, goodfile2)

    def tearDown(self):
        #self.driver.quit()
        pass        

if __name__ == '__main__':
    unittest.main()
